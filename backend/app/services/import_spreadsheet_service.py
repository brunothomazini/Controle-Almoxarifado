import os
import tempfile
from datetime import datetime
from typing import Optional
from pathlib import Path
from app.config import settings


class SpreadsheetImportService:
    def __init__(self, db):
        self.db = db

    def importar_relatorio(self, caminho_arquivo: str, tipo: str = "inventario") -> dict:
        ext = Path(caminho_arquivo).suffix.lower()

        if ext == ".xlsx":
            return self._importar_xlsx(caminho_arquivo, tipo)
        elif ext == ".csv":
            return self._importar_csv(caminho_arquivo, tipo)
        else:
            return {"erro": f"Formato nao suportado: {ext}. Use .xlsx ou .csv"}

    def _importar_xlsx(self, caminho: str, tipo: str) -> dict:
        try:
            import openpyxl
        except ImportError:
            return {"erro": "openpyxl nao instalado"}

        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        resultados = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            linhas = list(ws.iter_rows(values_only=True))
            if not linhas:
                continue

            cabecalho = [str(c or "").strip().upper() for c in linhas[0]]
            dados = linhas[1:]

            qtd = self._processar_dados(cabecalho, dados, sheet_name, tipo)
            resultados.append({"aba": sheet_name, "itens": qtd})

        wb.close()
        total = sum(r["itens"] for r in resultados)
        return {"status": "sucesso", "total_itens": total, "detalhes": resultados}

    def _importar_csv(self, caminho: str, tipo: str) -> dict:
        import csv

        with open(caminho, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            linhas = list(reader)

        if not linhas:
            return {"erro": "Arquivo CSV vazio"}

        cabecalho = [str(c or "").strip().upper() for c in linhas[0]]
        dados = linhas[1:]

        qtd = self._processar_dados(cabecalho, dados, "CSV", tipo)
        return {"status": "sucesso", "total_itens": qtd, "detalhes": [{"aba": "CSV", "itens": qtd}]}

    def _processar_dados(self, cabecalho: list, dados: list, sheet_name: str, tipo: str) -> int:
        from app.models.models import Item, Categoria, Fornecedor, Movimentacao, SyncLog

        col_map = {str(i).upper(): idx for idx, i in enumerate(cabecalho)}

        def _first(cols):
            for c in cols:
                v = col_map.get(c)
                if v is not None:
                    return v
            return None

        codigo_col = _first(["CODIGO", "CÓDIGO", "COD", "BEM", "ITEM", "ID"])
        nome_col = _first(["NOME COMERCIAL", "NOME", "DESCRICAO", "DESCRIÇÃO"])
        unidade_col = _first(["UNIDADE", "UNID"])
        quantidade_col = _first(["QUANTIDADE", "QTD", "ESTOQUE ATUAL", "SALDO"])
        grupo_col = _first(["GRUPO", "CATEGORIA", "TIPO", "CLASSE", "DEPARTAMENTO"])
        categoria_nome = sheet_name

        if codigo_col is None:
            codigo_col = col_map.get("ITEM")
        if codigo_col is None:
            codigo_col = col_map.get("ID")

        importados = 0

        for row in dados:
            try:
                if not row or all(v is None or str(v).strip() == "" for v in row):
                    continue

                codigo = str(row[codigo_col]).strip() if codigo_col is not None and codigo_col < len(row) and row[codigo_col] is not None else ""

                if not codigo:
                    continue

                nome = str(row[nome_col]).strip() if nome_col is not None and nome_col < len(row) and row[nome_col] is not None else ""
                unidade = str(row[unidade_col]).strip() if unidade_col is not None and unidade_col < len(row) and row[unidade_col] is not None else "UN"
                qtd = self._parse_float(row[quantidade_col]) if quantidade_col is not None and quantidade_col < len(row) else 0

                cat_nome = categoria_nome if grupo_col is None else str(row[grupo_col]).strip() if grupo_col is not None and grupo_col < len(row) and row[grupo_col] is not None else categoria_nome

                if tipo == "inventario":
                    self._atualizar_ou_criar_item(codigo, nome, unidade, qtd, cat_nome)

                elif tipo == "movimentacoes":
                    data_col = _first(["DATA", "DATA MOV", "DATA MOVIMENTAÇÃO"])
                    tipo_col = _first(["TIPO", "TIPO MOV", "ENTRADA/SAIDA"])
                    obs_col = _first(["OBSERVAÇÕES", "OBSERVACOES", "OBS", "HISTÓRICO", "HISTORICO"])

                    data_str = str(row[data_col]).strip() if data_col is not None and data_col < len(row) and row[data_col] is not None else ""
                    tipo_mov = str(row[tipo_col]).strip().lower() if tipo_col is not None and tipo_col < len(row) and row[tipo_col] is not None else "entrada"
                    obs = str(row[obs_col]).strip() if obs_col is not None and obs_col < len(row) and row[obs_col] is not None else ""

                    self._criar_movimentacao(codigo, qtd, tipo_mov, data_str, obs)

                importados += 1
            except Exception:
                continue

        self.db.commit()
        return importados

    def _atualizar_ou_criar_item(self, codigo: str, nome: str, unidade: str, quantidade: float, categoria_nome: str):
        from app.models.models import Item, Categoria

        cat = self.db.query(Categoria).filter(Categoria.nome == categoria_nome).first()
        if not cat:
            cat = Categoria(nome=categoria_nome)
            self.db.add(cat)
            self.db.flush()

        item = self.db.query(Item).filter(Item.codigo == codigo).first()
        if item:
            item.quantidade_atual = quantidade
            if nome:
                item.nome = nome
        else:
            item = Item(
                codigo=codigo,
                nome=nome or f"Item {codigo}",
                unidade_medida=unidade,
                quantidade_atual=quantidade,
                categoria_id=cat.id,
            )
            self.db.add(item)

    def _criar_movimentacao(self, codigo: str, quantidade: float, tipo: str, data_str: str, observacao: str):
        from app.models.models import Item, Movimentacao

        item = self.db.query(Item).filter(Item.codigo == codigo).first()
        if not item:
            return

        if tipo not in ("entrada", "saida"):
            tipo = "entrada"

        data = datetime.now()
        if data_str:
            try:
                if "-" in data_str and data_str.count("-") == 2:
                    data = datetime.strptime(data_str[:10], "%Y-%m-%d")
                elif "/" in data_str and data_str.count("/") == 2:
                    data = datetime.strptime(data_str[:10], "%d/%m/%Y")
            except (ValueError, IndexError):
                pass

        mov = Movimentacao(
            item_id=item.id,
            tipo=tipo,
            quantidade=quantidade,
            data=data,
            origem="planilha",
            observacoes=observacao or "Importado de relatorio",
        )
        self.db.add(mov)

        if tipo == "entrada":
            item.quantidade_atual = (item.quantidade_atual or 0) + quantidade
        else:
            item.quantidade_atual = max(0, (item.quantidade_atual or 0) - quantidade)

    def _parse_float(self, valor) -> float:
        if valor is None:
            return 0
        if isinstance(valor, (int, float)):
            return float(valor)
        try:
            return float(str(valor).replace(".", "").replace(",", ".").strip())
        except (ValueError, TypeError):
            return 0

    def salvar_arquivo(self, arquivo, nome_original: str) -> str:
        upload_dir = settings.UPLOAD_DIR / "relatorios"
        upload_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"{timestamp}_{nome_original}"
        caminho = str(upload_dir / nome_arquivo)

        with open(caminho, "wb") as f:
            content = arquivo.read()
            if isinstance(content, bytes):
                f.write(content)
            else:
                f.write(content.encode())

        return caminho
