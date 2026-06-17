import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import openpyxl
from app.database import SessionLocal
from app.models.models import Categoria, Item

XLSX = r"C:\Users\3580188\Documents\opencode\controle almoxarifado\Catalogo_de_Almoxarifado-itens-padronizados para dashboard.xlsx"

wb = openpyxl.load_workbook(XLSX, data_only=True)

db = SessionLocal()

item_map = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    bem_col = None
    first_data_row = 4

    for r in range(1, 6):
        for c in range(1, min(ws.max_column + 1, 10)):
            v = str(ws.cell(r, c).value or '').strip().upper()
            if v == 'BEM':
                bem_col = c
                first_data_row = r + 1
                break
        if bem_col:
            break

    if sheet_name == 'Informática':
        bem_col = 1
        first_data_row = 3

    if not bem_col:
        continue

    for r in range(first_data_row, ws.max_row + 1):
        val = ws.cell(r, bem_col).value
        if val is None:
            continue
        cod = str(val).strip()
        if not cod:
            continue
        item_map.setdefault(sheet_name, []).append(cod)

print("=== Categorias e quantidades de itens no Excel ===")
for cat, codes in sorted(item_map.items()):
    print(f"  {cat}: {len(codes)} itens")

print("\n=== Criando/obtendo categorias no banco ===")
cat_obj_map = {}
for sheet_name in item_map:
    existing = db.query(Categoria).filter(Categoria.nome == sheet_name).first()
    if existing:
        cat_obj_map[sheet_name] = existing
        print(f"  [EXISTE] {sheet_name} (id={existing.id})")
    else:
        nova = Categoria(nome=sheet_name)
        db.add(nova)
        db.flush()
        cat_obj_map[sheet_name] = nova
        print(f"  [CRIADA] {sheet_name} (id={nova.id})")

db.commit()

print("\n=== Atualizando itens ===")
total_atualizados = 0
nao_encontrados = []
ja_corretos = 0

for sheet_name, codes in item_map.items():
    cat_obj = cat_obj_map[sheet_name]
    for cod in codes:
        item = db.query(Item).filter(Item.codigo == cod).first()
        if not item:
            nao_encontrados.append(cod)
            continue
        if item.categoria_id == cat_obj.id:
            ja_corretos += 1
            continue
        item.categoria_id = cat_obj.id
        total_atualizados += 1

db.commit()

print(f"  Atualizados: {total_atualizados}")
print(f"  Ja estavam corretos: {ja_corretos}")
if nao_encontrados:
    print(f"  Nao encontrados no banco: {len(nao_encontrados)}")
    for c in nao_encontrados[:20]:
        print(f"    - {c}")
    if len(nao_encontrados) > 20:
        print(f"    ... e mais {len(nao_encontrados) - 20}")

db.close()
print("\nConcluido!")
